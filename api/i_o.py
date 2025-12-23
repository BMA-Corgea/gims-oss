# api/i_o.py

from __future__ import annotations
import io
import builtins
# These are the S3-aware functions we must use
from api.json_proxy import _is_s3_path, read_text, write_text, S3_ENABLED
import json
from pathlib import Path
from typing import Optional, Union, Any, Iterable, Iterator, Tuple, List, Callable
from api.manifest.resolver import resolve_path
import csv
from openpyxl import load_workbook
import zipfile
import mimetypes
import re

# ──────────────────────────────────────────────────────────────────────────────
# Debug
# ──────────────────────────────────────────────────────────────────────────────
DEBUG_ENABLED = False

def debug(*args, **kwargs):
    if DEBUG_ENABLED:
        print("[i_o]", *args, **kwargs)

# -----------------------------
# Load full schemas from file
# -----------------------------

def load_schema(project_path: Path, word_type: str) -> dict:
    """
    Load the full *_types.json schema file. (S3-AWARE)
    word_type: one of 'noun', 'verb', 'adjective', 'adverb'
    """
    schema_path = project_path / f"{word_type}_types.json"
    try:
        payload = read_text(schema_path, encoding="utf-8")
        return json.loads(payload)
    except Exception as e:
        msg = str(e)
        # Catch S3 "NoSuchKey" or local "FileNotFoundError"
        if "NoSuchKey" in msg or isinstance(e, FileNotFoundError):
            debug(f"[load_schema] {schema_path} not found (local or S3).")
            raise FileNotFoundError(f"{schema_path} not found.")
        else:
            # A different error (e.g., invalid JSON), let it crash
            debug(f"[load_schema] Failed to load/parse schema {schema_path}", {"error": repr(e)})
            raise e


def load_override(project_path: Path) -> list[dict]:
    """
    Load the override.json file (list of override instructions). (S3-AWARE)
    """
    override_path = project_path / "override.json"
    try:
        payload = read_text(override_path, encoding="utf-8")
        return json.loads(payload)
    except FileNotFoundError:
        return []  # Original behavior
    except Exception:
        return []

def load_autogen_counters(project_path: Path) -> dict:
    """
    Load autogen_counters.json — used to track unique ID segment counters. (S3-AWARE)
    Returns an empty dict if the file doesn't exist.
    """
    path = project_path / "autogen_counters.json"
    try:
        payload = read_text(path, encoding="utf-8")
        return json.loads(payload)
    except FileNotFoundError:
        return {}
    except Exception:
        return {}

# -----------------------------
# Schema lookup by top-level key
# -----------------------------

def get_noun_schema(project_path: Path, noun_name: str) -> Optional[dict]:
    schema = load_schema(project_path, "noun")
    return schema.get(noun_name)


def get_verb_schema(project_path: Path, verb_name: str) -> Optional[dict]:
    schema = load_schema(project_path, "verb")
    return schema.get(verb_name)


def get_adjective_schema(
    project_path: Path,
    adjective_name: str,
    applies_to: Optional[str] = None
) -> Optional[dict]:
    schema_list = load_schema(project_path, "adjective")
    candidates = [
        entry for entry in schema_list
        if entry.get("adjective") == adjective_name
    ]

    if applies_to:
        for entry in candidates:
            if applies_to in entry.get("applies_to", []):
                return entry
        return None
    else:
        return candidates[0] if candidates else None


def get_adverb_schema(
    project_path: Path,
    adverb_name: str,
    applies_to: Optional[str] = None
) -> Optional[dict]:
    schema_list = load_schema(project_path, "adverb")

    candidates = [
        entry for entry in schema_list
        if entry.get("adverb") == adverb_name
    ]

    if applies_to:
        for entry in candidates:
            if applies_to in entry.get("applies_to", []):
                return entry
        return None
    else:
        return candidates[0] if candidates else None

def get_autogen_counter(project_path: Path, noun_type: str) -> Optional[dict]:
    all_counters = load_autogen_counters(project_path)
    return all_counters.get(noun_type)


def get_override_schema(project_path: Path, run_id: str) -> list[dict]:
    overrides = load_override(project_path)
    return [entry for entry in overrides if entry.get("run") == run_id]

# -----------------------------
# Schema lookup by not-top-level key
# -----------------------------

def find_non_id_field_value(
    project_path: Path,
    search_value: str,
    word_type: str | list[str] | None = None
) -> list[dict]:
    """
    Search one or more *_types.json files for entries where a non-identifying field
    has a value that matches `search_value`.
    """
    def walk(obj, skip_keys: set, path=""):
        matches = []
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k in skip_keys:
                    continue
                new_path = f"{path}.{k}" if path else k
                matches += walk(v, skip_keys, new_path)
        elif isinstance(obj, list):
            for i, v in enumerate(obj):
                new_path = f"{path}[{i}]"
                matches += walk(v, skip_keys, new_path)
        elif obj == search_value:
            matches.append(path)
        return matches

    results = []

    valid_types = {"noun", "verb", "adjective", "adverb"}
    types_to_search = (
        [word_type] if isinstance(word_type, str)
        else word_type if isinstance(word_type, list)
        else ["noun", "verb", "adjective", "adverb"]
    )

    for wt in types_to_search:
        if wt not in valid_types:
            continue

        try:
            schema = load_schema(project_path, wt)
        except Exception:
            continue

        if isinstance(schema, dict):  # noun_types or verb_types
            for schema_name, entry in schema.items():
                matches = walk(entry, skip_keys=set())
                for path in matches:
                    results.append({
                        "word_type": wt,
                        "schema_name": schema_name,
                        "match_path": path,
                        "matched_value": search_value,
                        "schema": entry
                    })

        elif isinstance(schema, list):  # adjective_types or adverb_types
            id_field = "adjective" if wt == "adjective" else "adverb"
            for entry in schema:
                schema_name = entry.get(id_field, "(unknown)")
                matches = walk(entry, skip_keys={id_field})
                for path in matches:
                    results.append({
                        "word_type": wt,
                        "schema_name": schema_name,
                        "match_path": path,
                        "matched_value": search_value,
                        "schema": entry
                    })

    return results

def find_in_override_by_non_id_field_value(project_path: Path, search_value: str) -> list[dict]:
    """
    Search override.json for any entries where a non-identifying field (not 'run') 
    has a value matching `search_value`. (S3-AWARE)
    """
    path = project_path / "override.json"
    try:
        payload = read_text(path, encoding="utf-8")
        entries = json.loads(payload)
    except Exception:
        return []

    results = []
    for entry in entries:
        for key, val in entry.items():
            if key == "run":
                continue
            if isinstance(val, str) and val == search_value:
                results.append({
                    "match_path": key,
                    "matched_value": val,
                    "entry": entry
                })
            elif isinstance(val, list) and search_value in val:
                idx = val.index(search_value)
                results.append({
                    "match_path": f"{key}[{idx}]",
                    "matched_value": search_value,
                    "entry": entry
                })
    return results

# -----------------------------
# Write Operations (JSON / JSONL)
# -----------------------------

def save_schema(project_path: Path, word_type: str, data: dict):
    """
    Overwrite the full *_types.json file for the given word type. (S3-AWARE)
    """
    path = project_path / f"{word_type}_types.json"
    write_text(path, json.dumps(data, indent=2), encoding="utf-8")


def save_autogen_counters(project_path: Path, data: dict):
    path = project_path / "autogen_counters.json"
    write_text(path, json.dumps(data, indent=2), encoding="utf-8")


def save_override(project_path: Path, data: list[dict]):
    path = project_path / "override.json"
    write_text(path, json.dumps(data, indent=2), encoding="utf-8")


def append_jsonl(path: Path, entry: dict):
    """
    Append a new JSON line to a .jsonl file. (S3-AWARE)
    """
    try:
        payload = read_text(path, encoding="utf-8")
        lines = payload.splitlines()
    except FileNotFoundError:
        lines = []
    
    lines.append(json.dumps(entry))
    new_payload = "\n".join(lines) + "\n"
    write_text(path, new_payload, encoding="utf-8")


def replace_jsonl_entry(path: Path, match: Callable[[dict], bool], new_entry: dict):
    """
    Replace a single entry in a .jsonl file where match(entry) is True. (S3-AWARE)
    """
    try:
        payload = read_text(path, encoding="utf-8")
        lines = [json.loads(line) for line in payload.splitlines() if line.strip()]
    except FileNotFoundError:
        raise ValueError("File not found, cannot replace entry.")

    updated = False
    for i, line in enumerate(lines):
        if match(line):
            lines[i] = new_entry
            updated = True
            break
    if not updated:
        raise ValueError("Entry not found for replacement.")
    
    new_payload = "".join(json.dumps(line) + "\n" for line in lines)
    write_text(path, new_payload, encoding="utf-8")

def rewrite_jsonl(path: Path, entries: list[dict]):
    """
    Atomically rewrite a .jsonl file with the given list of dicts. (S3-AWARE)
    """
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    
    payload = "".join(json.dumps(e) + "\n" for e in entries)
    write_text(path, payload, encoding="utf-8")

def save_json(path: Path, data: dict):
    """
    Save a full dict to a .json file (e.g., DataEntry.json, Status.json). (S3-AWARE)
    """
    write_text(path, json.dumps(data, indent=2), encoding="utf-8")

# Small convenience wrappers (seen missing in logs previously)
def read_json(path: Path, default=None):
    try:
        txt = read_text(path, encoding="utf-8")
        return json.loads(txt)
    except FileNotFoundError:
        return default
    except Exception:
        return default

def write_json(path: Path, data: Any):
    save_json(path, data)

# -----------------------------
# Verb Group Logs (SQL + RDS-aware; with JSONL fallback)
# -----------------------------

try:
    import psycopg  # psycopg[binary]
    _PSYCOPG_AVAILABLE = True
except Exception:
    _PSYCOPG_AVAILABLE = False

import sqlite3  # used for local objects.db too

def _normalize_for_psycopg(url: str) -> str:
    """Normalize SQLAlchemy-style URIs to plain psycopg DSNs."""
    if url.startswith("postgresql+"):
        url = "postgresql://" + url.split("postgresql+", 1)[1]
    if "?ssl=require" in url:
        url = url.replace("?ssl=require", "?sslmode=require")
    return url.replace("postgresql://asyncpg://", "postgresql://")

def _get_objects_db_target(project_path: Path) -> tuple[str, str]:
    """
    Decide ('pg'|'sqlite', target) using resolver + RDS module, mirroring server behavior.
    """
    try:
        from api.manifest.resolver import get_db_uri
    except Exception:
        get_db_uri = None

    uri = None
    if get_db_uri:
        try:
            uri = get_db_uri("object_sql_db")
        except Exception:
            uri = None

    if uri and uri.startswith("postgresql"):
        return ("pg", _normalize_for_psycopg(uri))

    db_path = resolve_path(project_path, "object_sql_db")
    return ("sqlite", db_path.as_posix())

def _table_name(project: str) -> str:
    """Match the local convention: <project-with-hyphens>_verb_log (always quoted)."""
    return f"{project.replace('_','-')}_verb_log"

def get_verb_group_log_config(project_path: Path, verb_group: str) -> dict:
    """
    Load the verb-group log config using resolve_path.
    Map key: 'verb_group_log_config' => verbs/{verb_group}/{verb_group}_log_config.json
    (S3-AWARE via read_text)
    """
    cfg_path = resolve_path(project_path, "verb_group_log_config", verb_group=verb_group)
    try:
        payload = read_text(cfg_path, encoding="utf-8")
        return json.loads(payload) if payload else {}
    except FileNotFoundError:
        raise FileNotFoundError(f"{verb_group}_log_config.json not found for verb group '{verb_group}' at {cfg_path}")
    except Exception as e:
        debug(f"[get_verb_group_log_config] Failed to load {cfg_path}: {e!r}")
        raise

def _json_from_db_cell(cell):
    if isinstance(cell, (dict, list)):
        return cell
    try:
        return json.loads(cell)
    except Exception:
        return cell

def load_verb_group_log(project_path: Path, verb_group: str) -> list[dict]:
    """
    Load all entries for a verb group from SQL table, with S3-AWARE JSONL fallback.
    """
    kind, target = _get_objects_db_target(project_path)
    table = _table_name(project_path.name)

    try:
        if kind == "pg" and _PSYCOPG_AVAILABLE:
            with psycopg.connect(target) as conn, conn.cursor() as cur:
                cur.execute(f'SELECT data FROM public."{table}" WHERE verb_group=%s ORDER BY ts DESC;', (verb_group,))
                rows = cur.fetchall()
            return [_json_from_db_cell(r[0]) for r in rows]

        conn = sqlite3.connect(target)
        try:
            c = conn.cursor()
            c.execute(f'SELECT data FROM "{table}" WHERE verb_group=? ORDER BY ts DESC;', (verb_group,))
            rows = c.fetchall()
        finally:
            conn.close()
        return [_json_from_db_cell(r[0]) for r in rows]

    except Exception as e:
        debug(f"[load_verb_group_log] DB read failed -> JSONL fallback: {e}")

    # JSONL fallback (S3-AWARE)
    log_path = resolve_path(project_path, "verb_group_log", verb_group=verb_group)
    try:
        payload = read_text(log_path, encoding="utf-8")
        return [json.loads(line) for line in payload.splitlines() if line.strip()]
    except FileNotFoundError:
        return []

def append_to_verb_group_log(project_path: Path, verb_group: str, entry: dict):
    """
    Insert/upsert to SQL table, with S3-AWARE JSONL fallback.
    """
    cfg = get_verb_group_log_config(project_path, verb_group)
    primary_id_field = cfg.get("primary_id") or "run_id"
    pid_val = entry.get(primary_id_field)
    if not pid_val:
        raise ValueError(f"Entry missing primary ID field '{primary_id_field}'")
    verb = entry.get("test_type") or entry.get("verb")
    if not verb:
        raise ValueError("Entry missing 'test_type' or 'verb'")

    kind, target = _get_objects_db_target(project_path)
    table = _table_name(project_path.name)

    try:
        if kind == "pg" and _PSYCOPG_AVAILABLE:
            with psycopg.connect(target, autocommit=True) as conn, conn.cursor() as cur:
                cur.execute(
                    f"""
                    INSERT INTO public."{table}" (primary_id, verb_group, verb, data)
                    VALUES (%s, %s, %s, %s::jsonb)
                    ON CONFLICT (primary_id) DO UPDATE
                    SET verb_group=EXCLUDED.verb_group,
                        verb=EXCLUDED.verb,
                        data=EXCLUDED.data;
                    """,
                    (str(pid_val), verb_group, str(verb), json.dumps(entry)),
                )
            return

        conn = sqlite3.connect(target)
        try:
            c = conn.cursor()
            c.execute(
                f'UPDATE "{table}" SET verb_group=?, verb=?, data=? WHERE primary_id=?',
                (verb_group, str(verb), json.dumps(entry), str(pid_val)),
            )
            if c.rowcount == 0:
                c.execute(
                    f'INSERT INTO "{table}" (primary_id, verb_group, verb, data) VALUES (?, ?, ?, ?)',
                    (str(pid_val), verb_group, str(verb), json.dumps(entry)),
                )
            conn.commit()
        finally:
            conn.close()
        return

    except Exception as e:
        debug(f"[append_to_verb_group_log] DB write failed -> JSONL fallback: {e}")

    # JSONL fallback (S3-AWARE)
    log_path = resolve_path(project_path, "verb_group_log", verb_group=verb_group)
    append_jsonl(log_path, entry)

def replace_in_verb_group_log(project_path: Path, verb_group: str, new_entry: dict):
    """
    Replace in SQL table, with S3-AWARE JSONL fallback.
    """
    cfg = get_verb_group_log_config(project_path, verb_group)
    primary_id_field = cfg.get("primary_id") or "run_id"
    pid_val = new_entry.get(primary_id_field)
    if not pid_val:
        raise ValueError(f"New entry missing primary ID field '{primary_id_field}'")
    verb = new_entry.get("test_type") or new_entry.get("verb")
    if not verb:
        raise ValueError("Entry missing 'test_type' or 'verb'")

    kind, target = _get_objects_db_target(project_path)
    table = _table_name(project_path.name)

    try:
        if kind == "pg" and _PSYCOPG_AVAILABLE:
            with psycopg.connect(target, autocommit=True) as conn, conn.cursor() as cur:
                cur.execute(
                    f'UPDATE public."{table}" SET verb_group=%s, verb=%s, data=%s::jsonb WHERE primary_id=%s;',
                    (verb_group, str(verb), json.dumps(new_entry), str(pid_val)),
                )
                if cur.rowcount == 0:
                    raise ValueError(f"No entry found with primary_id='{pid_val}'")
            return

        conn = sqlite3.connect(target)
        try:
            c = conn.cursor()
            c.execute(
                f'UPDATE "{table}" SET verb_group=?, verb=?, data=? WHERE primary_id=?',
                (verb_group, str(verb), json.dumps(new_entry), str(pid_val)),
            )
            if c.rowcount == 0:
                raise ValueError(f"No entry found with primary_id='{pid_val}'")
            conn.commit()
        finally:
            conn.close()
        return

    except Exception as e:
        debug(f"[replace_in_verb_group_log] DB update failed -> JSONL fallback: {e}")

    log_path = resolve_path(project_path, "verb_group_log", verb_group=verb_group)

    def match(entry: dict):
        return entry.get(primary_id_field) == pid_val
    
    replace_jsonl_entry(log_path, match, new_entry)

# -----------------------------
# Noun + Verb file references
# -----------------------------

def resolve_verb_group_from_test_type(project_path: Path, test_type: str) -> str | None:
    schema = load_schema(project_path, "verb")
    verb_schema = schema.get(test_type)
    if not verb_schema:
        return None
    return verb_schema.get("verb_group", "Tests")

def resolve_reference_noun_from_verb(project_path: Path, verb_name: str) -> Optional[str]:
    verb_schema = get_verb_schema(project_path, verb_name)
    if not verb_schema:
        return None

    return (
        verb_schema.get("data_entry_schema", {})
        .get("set_up_inputs", {})
        .get("noun_type_ref")
    )

def get_noun_items(project_path: Path, noun_type: str) -> list[dict]:
    """
    Load entries for a noun (RDS + SQLite + S3-AWARE JSONL fallback).
    """
    import sqlite3
    sanitized = re.sub(r"\W+", "_", str(noun_type)).strip("_")
    base_table = f"noun_{sanitized}"
    project_name = project_path.name
    full_table = f"{project_name.replace('_','-')}_{base_table}"
    kind, target = _get_objects_db_target(project_path)
    debug(f"[get_noun_items] noun={noun_type!r} kind={kind} target={target}")

    def _normalize_row(d: dict) -> dict:
        fixed = dict(d)
        for k in list(d.keys()):
            if "_" in k:
                spaced = k.replace("_", " ")
                fixed.setdefault(spaced, d[k])
            if " " in k:
                underscored = k.replace(" ", "_")
                fixed.setdefault(underscored, d[k])
        return fixed

    # 1) RDS
    if kind == "pg" and _PSYCOPG_AVAILABLE:
        try:
            with psycopg.connect(target) as conn, conn.cursor() as cur:
                # Check existence
                cur.execute("SELECT to_regclass(%s);", (f'public."{full_table}"',))
                exists = cur.fetchone()[0]
                if not exists:
                    debug(f"[get_noun_items] X RDS table '{full_table}' not found.")
                    raise Exception("no_table")

                # Fetch all rows now that we know it exists
                cur.execute(f'SELECT * FROM public."{full_table}" ORDER BY 1;')
                colnames = [desc.name for desc in cur.description]
                rows = cur.fetchall()

            out = [_normalize_row(dict(zip(colnames, r))) for r in rows]
            debug(f"[get_noun_items] ✓ returning {len(out)} rows from RDS table {full_table}")
            return out
        except Exception as e:
            debug(f"[get_noun_items] RDS read failed -> fallback: {e}")

    # 2) SQLite
    if kind == "sqlite":
        try:
            conn = sqlite3.connect(target)
            conn.row_factory = sqlite3.Row
            try:
                c = conn.cursor()
                table_names = [full_table, base_table]
                for table in table_names:
                    exists = c.execute(
                        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                        (table,)
                    ).fetchone()
                    if exists:
                        rows = c.execute(f'SELECT * FROM "{table}"').fetchall()
                        out = [_normalize_row(dict(r)) for r in rows]
                        debug(f"[get_noun_items] ✓ returning {len(out)} rows from SQLite table {table}")
                        return out
                debug(f"[get_noun_items] no SQLite table found for {noun_type}")
            finally:
                conn.close()
        except Exception as e:
            debug(f"[get_noun_items] SQLite read failed -> fallback: {e}")

    # 3) JSONL fallback
    jsonl_path = resolve_path(project_path, "noun_items", noun_type=noun_type)
    try:
        payload = read_text(jsonl_path, encoding="utf-8")
        items = [json.loads(line) for line in payload.splitlines() if line.strip()]
    except FileNotFoundError:
        debug(f"[get_noun_items] X no SQL and no JSONL -> []")
        return []

    out = [_normalize_row(i) if isinstance(i, dict) else i for i in items]
    debug(f"[get_noun_items] ✓ returning {len(out)} rows (JSONL fallback)")
    return out


# In i_o.py, inside get_noun_items()
    # 3) JSONL fallback
    jsonl_path = resolve_path(project_path, "noun_items", noun_type=noun_type)
    try:
        payload = read_text(jsonl_path, encoding="utf-8")
        items = [json.loads(line) for line in payload.splitlines() if line.strip()]
    except Exception as e:
        msg = str(e)
        # This now catches FileNotFoundError AND S3 "NoSuchKey" errors
        if "NoSuchKey" in msg or isinstance(e, FileNotFoundError):
            debug(f"[get_noun_items] X no SQL and no JSONL -> []")
            return []
        else:
            # A different error (e.g., invalid JSON), let it crash
            debug(f"[get_noun_items] ERROR reading JSONL: {e!r}")
            raise e

    out = [_normalize_row(i) if isinstance(i, dict) else i for i in items]
    debug(f"[get_noun_items] ✓ returning {len(out)} rows (JSONL fallback)")
    return out

def resolve_noun_type_from_override(project_path: Path, override: dict) -> str | None:
    """
    Return noun_type for an override entry via verb schema.
    """
    run_id = override.get("run")
    verb = override.get("verb")
    if not run_id or not verb:
        return None
    try:
        verb_schema = get_verb_schema(project_path, verb)
    except Exception:
        return None
    noun_type = (
        verb_schema.get("data_entry_schema", {})
        .get("set_up_inputs", {})
        .get("noun_type_ref")
    )
    return noun_type  # <- FIXED: previously missing return

def resolve_run_id_to_test_type(project_path: Path, run_id: str) -> str:
    """
    Resolve a run_id -> test_type (or 'verb') by scanning verb-group logs.
    Uses resolve_path-backed helpers (list_verb_groups + load_verb_group_log)
    and each group's {verb_group}_log_config to determine the primary id field.
    (S3/RDS-aware through the helpers)
    """
    # Iterate all verb groups discovered via resolver-backed listing
    groups: list[str] = []
    try:
        groups = list_verb_groups(project_path)
    except Exception as e:
        debug(f"[resolve_run_id_to_test_type] list_verb_groups failed: {e!r}")
        groups = []

    for group in groups:
        # Determine primary id key from the group's config (fallbacks included)
        try:
            cfg = get_verb_group_log_config(project_path, group)
            primary_id = (
                cfg.get("primary_id")
                or cfg.get("primaryId")
                or cfg.get("primary_id_field")
                or "run_id"  # sensible default
            )
        except Exception:
            primary_id = "run_id"

        # Build a small set of candidate keys to tolerate space/underscore variants
        cand_keys = {
            primary_id,
            primary_id.replace("_", " "),
            primary_id.replace(" ", "_"),
            "run_id", "run_ID", "Run ID"  # extra safety nets for legacy rows
        }

        # Load entries from SQL (RDS/SQLite) with JSONL fallback via resolver
        try:
            entries = load_verb_group_log(project_path, group)
        except Exception as e:
            debug(f"[resolve_run_id_to_test_type] load_verb_group_log({group}) failed: {e!r}")
            entries = []

        # Scan rows for a matching run id
        for row in entries:
            if not isinstance(row, dict):
                continue
            # quick normalization pass to allow both spaced and underscored access
            # without mutating original dict
            norm = dict(row)
            for k, v in list(row.items()):
                if isinstance(k, str):
                    if " " in k:
                        norm.setdefault(k.replace(" ", "_"), v)
                    if "_" in k:
                        norm.setdefault(k.replace("_", " "), v)

            # compare as strings to avoid type surprises
            for k in cand_keys:
                if k in norm and str(norm.get(k)) == str(run_id):
                    verb_name = norm.get("test_type") or norm.get("verb")
                    if verb_name:
                        return str(verb_name)

    # Nothing matched
    raise ValueError(f"Run ID {run_id!r} not found in any verb-group log.")


def list_verb_groups(project_path: Path) -> list[str]:
    """
    S3-aware listing of verb groups.
    """
    verbs_dir = resolve_path(project_path, "verbs_dir")
    if not fs_exists(verbs_dir) or not fs_is_dir(verbs_dir):
        raise FileNotFoundError(f"X verbs_dir not found at {verbs_dir}")
    return [Path(d).name for d in fs_iterdir(verbs_dir) if fs_is_dir(d)]

# -----------------------------
# Generic load
# -----------------------------

def load_data(path: Path, *, default=None, strict: bool = False, encoding: str = "utf-8"):
    """
    Read JSON (or text if .md/.csv) from `path`. On missing file in S3/FS:
      - if strict=True → raise
      - else → return `default`
    """
    try:
        text = read_text(path, encoding=encoding)  # S3-aware
    except Exception as e:
        msg = str(e)
        not_found = (
            isinstance(e, FileNotFoundError)
            or "NoSuchKey" in msg
            or "The specified key does not exist" in msg
            or "Not Found" in msg
        )
        if not strict and not_found:
            return default
        raise ValueError(f"X Failed to load data from {path}: {e}")

    low = path.suffix.lower()
    if low == ".json":
        return json.loads(text) if text else default
    if low == ".jsonl":
        # Return list for JSONL
        return [json.loads(line) for line in text.splitlines() if line.strip()] if text else default
    if low in {".md", ".csv", ".txt", ""}:
        return text if text is not None else default
    return json.loads(text) if text else default


def load_table_data(path: Path) -> list[dict[str, Any]] | None:
    """
    Load a table-like file (CSV or XLSX) and return as a list of dictionaries. (S3-AWARE)
    """
    if path.suffix.lower() == ".csv":
        try:
            with open_file(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                rows = [row for row in reader if any((cell or "").strip() for cell in row.values())]
                return rows or None
        except FileNotFoundError:
            return None
        except Exception as e:
            raise ValueError(f"X Failed to load CSV from {path}: {e}")

    elif path.suffix.lower() == ".xlsx":
        try:
            with open_file(path, "rb") as f:
                wb = load_workbook(f, read_only=True, data_only=True)
                sheet = wb.active
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                if not any(headers):
                    return None
                rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    if any(str(cell).strip() for cell in row_dict.values() if cell is not None):
                        rows.append(row_dict)
                return rows or None
        except FileNotFoundError:
            return None
        except Exception as e:
            debug(f"Failed to load XLSX, may be S3 binary issue: {e}")
            raise ValueError(f"X Failed to load XLSX from {path}: {e}")

    else:
        raise ValueError(f"X Unsupported file type for table loading: {path.suffix}")


def is_file_empty(path: Path) -> bool:
    """
    S3-aware emptiness check.
    - For text-like types, read via read_text (S3-aware).
    - For CSV, stream via open_file (S3-aware).
    - For binaries (xlsx/docx/pdf/zip/images), stream via fs_open_readbin (S3-aware).
    - Falls back to size/stat checks using fs_stat_size for S3 and local stat() for FS.
    """
    ext = path.suffix.lower()

    try:
        # --- Text-ish files: read as text (works for S3 + local)
        if ext in {".txt", ".log", ".md", ".html", ".xml", ".json"}:
            return not (read_text(path, encoding="utf-8", errors="ignore").strip())

        # --- CSV: iterate rows (S3 + local via open_file)
        if ext == ".csv":
            with open_file(path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                return not any(row for row in reader)

        # --- Quick size check (S3 uses fs_stat_size; local uses stat)
        if S3_ENABLED and _is_s3_path(path):
            if fs_stat_size(path) == 0:
                return True
        else:
            if (not path.exists()) or path.stat().st_size == 0:
                return True

        # --- XLSX: check for any non-empty cell (S3 + local)
        if ext == ".xlsx":
            import openpyxl
            with fs_open_readbin(path) as fh:
                wb = openpyxl.load_workbook(fh, data_only=True)
            return all(
                not any(cell.value for row in sheet.iter_rows() for cell in row)
                for sheet in wb.worksheets
            )

        # --- DOCX: check for any non-empty paragraph (S3 + local)
        if ext == ".docx":
            import docx
            with fs_open_readbin(path) as fh:
                doc = docx.Document(fh)
            return not any((p.text or "").strip() for p in doc.paragraphs)

        # --- PDF: any page with extractable text? (S3 + local)
        if ext == ".pdf":
            import PyPDF2
            with fs_open_readbin(path) as fh:
                reader = PyPDF2.PdfReader(fh)
                return all(not ((page.extract_text() or "").strip()) for page in reader.pages)

        # --- ZIP: all members empty or only directories? (S3 + local)
        # (We skip zipfile.is_zipfile(path) to be S3-safe and inspect via ZipFile directly.)
        try:
            with fs_open_readbin(path) as fh:
                with zipfile.ZipFile(fh) as z:
                    return all(
                        info.file_size == 0
                        for info in z.infolist()
                        if not info.filename.endswith("/")
                    )
        except zipfile.BadZipFile:
            # Not a zip; continue to image/mime handling below
            pass

        # --- Images: check for 0x0 size (S3 + local)
        if ext in {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff"}:
            from PIL import Image
            with fs_open_readbin(path) as fh:
                with Image.open(fh) as img:
                    return img.size == (0, 0)

        # --- Fallback mime-based text sniff
        mime_type, _ = mimetypes.guess_type(path.name)
        if mime_type and mime_type.startswith("text"):
            return not (read_text(path, encoding="utf-8", errors="ignore").strip())

        # If we got here, treat as non-empty by default (binary of unknown type)
        return False

    except FileNotFoundError:
        return True
    except Exception as e:
        debug(f"! Error checking {path.name}: {e}")
        return False


def load_local_layout_map(project_path: Path) -> dict:
    """
    (Unchanged) This file MUST be local to configure the S3/RDS connection.
    """
    here = Path(__file__).resolve()
    repo_root = here.parent
    fpath = repo_root / "manifest" / "local_layout_map.json"
    if not fpath.exists():
        raise FileNotFoundError(f"local_layout_map.json not found at {fpath}")

    data = json.loads(fpath.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        if "db_map" in data and isinstance(data["db_map"], dict):
            return data["db_map"]
        if "db_endpoints" in data and isinstance(data["db_endpoints"], dict):
            return data["db_endpoints"]
    return data if isinstance(data, dict) else {}

def _sanitize_table_name(noun: str) -> str:
    base = re.sub(r"[^0-9a-zA-Z_]", "_", noun).strip("_")
    if not base or not base[0].isalpha():
        base = f"T_{base}"
    return f"noun_{base}"

def get_url_base(project_path: Path) -> str:
    """
    Return the base API URL for this project. (S3-AWARE)
    Looks for <project>/url_base.txt. If missing, fall back to localhost:8000.
    """
    f = project_path / "url_base.txt"
    try:
        return read_text(f, encoding="utf-8").strip()
    except FileNotFoundError:
        return "http://127.0.0.1:8000"

def open_file(path, mode="r", encoding="utf-8", **kwargs):
    """
    Unified file opener for local and S3 paths.
    Accepts extra **kwargs (e.g., errors="ignore") to mirror builtins.open API.
    """
    errors = kwargs.get("errors", None)

    # ---- Local FS fast-path ---------------------------------------------------
    if not S3_ENABLED:
        if "b" in mode:
            return builtins.open(path, mode)
        return builtins.open(path, mode, encoding=encoding, errors=errors)

    # ---- S3 mode --------------------------------------------------------------
    try:
        from api import json_proxy as _jp  # type: ignore
    except Exception:
        _jp = None

    has_rb = bool(_jp and getattr(_jp, "read_bytes", None))
    has_wb = bool(_jp and getattr(_jp, "write_bytes", None))

    # BINARY MODES --------------------------------------------------------------
    if "b" in mode:
        # Read
        if "r" in mode and has_rb:
            data = _jp.read_bytes(path)  # type: ignore[attr-defined]
            return io.BytesIO(data)

        # Write / Append
        if ("w" in mode or "a" in mode) and has_wb:
            initial = b""
            if "a" in mode:
                try:
                    initial = _jp.read_bytes(path)  # type: ignore[attr-defined]
                except FileNotFoundError:
                    initial = b""

            buffer = io.BytesIO(initial)

            # Safer close monkey-patch that preserves the original close()
            orig_close = buffer.close
            def _close_and_upload_bin():
                body = buffer.getvalue()
                _jp.write_bytes(path, body)  # type: ignore[attr-defined]
                orig_close()
            buffer.close = _close_and_upload_bin  # type: ignore[assignment]

            return buffer

        # No byte helpers available → fall back to local FS (best-effort)
        debug("[i_o.open_file] WARNING: Binary mode without S3 byte helpers; using local FS.")
        return builtins.open(path, mode)

    # TEXT MODES ---------------------------------------------------------------
    if "r" in mode:
        # read_text is S3-aware
        data = read_text(path, encoding=encoding, errors=errors)
        return io.StringIO(data)

    if "w" in mode or "a" in mode:
        # Seed with existing content for append
        initial_data = ""
        if "a" in mode:
            try:
                initial_data = read_text(path, encoding=encoding)
                if not initial_data.endswith("\n"):
                    initial_data += "\n"
            except FileNotFoundError:
                pass

        buffer = io.StringIO(initial_data)
        if "a" in mode:
            buffer.seek(0, io.SEEK_END)

        # Safer close monkey-patch that preserves the original close()
        orig_close = buffer.close
        def _close_and_upload():
            body = buffer.getvalue()
            write_text(path, body, encoding=encoding)
            orig_close()
        buffer.close = _close_and_upload  # type: ignore[assignment]

        return buffer

    raise ValueError(f"Unsupported mode: {mode}")

# ──────────────────────────────────────────────────────────────────────────────
# [LOCAL_IO_SHIM]  S3-aware filesystem helper layer (+ project lister)
# ──────────────────────────────────────────────────────────────────────────────

def s3_capabilities() -> dict:
    """
    Probe optional json_proxy capabilities so callers can degrade gracefully.
    """
    caps = {
        "exists": False,
        "isfile": False,
        "isdir": False,
        "stat": False,
        "listdir": False,
        "walk": False,
        "read_bytes": False,
        "write_bytes": False,
        "delete": False,
        "list_projects": False,
        "list_dirnames": False,
        "presign": False,
        "iterdir": False,
    }
    try:
        from api import json_proxy as _jp  # type: ignore
    except Exception:
        return caps

    for name in list(caps.keys()):
        caps[name] = bool(getattr(_jp, name, None))
    if not caps["list_dirnames"]:
        caps["list_dirnames"] = bool(getattr(_jp, "list_children", None) or getattr(_jp, "list_prefixes", None))
    return caps

def _s3_call(name: str, *args, default=None, **kwargs):
    """
    Safe invoker for optional json_proxy functions. Returns default if missing/fails.
    """
    try:
        from api import json_proxy as _jp  # type: ignore
    except Exception:
        return default
    fn = getattr(_jp, name, None)
    if not fn:
        if name == "list_dirnames":
            fn = getattr(_jp, "list_children", None) or getattr(_jp, "list_prefixes", None)
        if name == "iterdir":
            fn = getattr(_jp, "iterdir", None)
    if not fn:
        return default
    try:
        return fn(*args, **kwargs)  # type: ignore[misc]
    except Exception as e:
        debug(f"[s3_call:{name}] {repr(e)}")
        return default

# ---- Path predicates ---------------------------------------------------------
#
# *** FIX 3: REPLACING THIS ENTIRE BLOCK OF FS SHIMS ***
#
def fs_exists(path: Path) -> bool:
    if not S3_ENABLED or not _is_s3_path(path):
        return path.exists()
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "exists"):
            return bool(_jp.exists(str(path)))
    except Exception as e:
        debug(f"[fs_exists] S3 call failed: {e!r}, falling back to local")
    return path.exists() # Fallback

def fs_is_file(path: Path) -> bool:
    if not S3_ENABLED or not _is_s3_path(path):
        return path.is_file()
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "is_file"):
            return bool(_jp.is_file(str(path)))
    except Exception as e:
        debug(f"[fs_is_file] S3 call failed: {e!r}, falling back to local")
    return path.is_file() # Fallback

def fs_is_dir(path: Path) -> bool:
    if not S3_ENABLED or not _is_s3_path(path):
        return path.is_dir()
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "is_dir"):
            return bool(_jp.is_dir(str(path)))
    except Exception as e:
        debug(f"[fs_is_dir] S3 call failed: {e!r}, falling back to local")
    return path.is_dir() # Fallback

# ---- Stat / size -------------------------------------------------------------

def fs_makedirs(path: str | Path, exist_ok: bool = True) -> None:
    """
    Create a directory path, S3-aware. Uses json_proxy.fs_makedirs when available,
    else falls back to os.makedirs for local filesystem.
    """
    import os
    path = str(path)
    try:
        from api import json_proxy
    except Exception:
        json_proxy = None

    if S3_ENABLED and json_proxy and hasattr(json_proxy, "fs_makedirs"):
        try:
            json_proxy.fs_makedirs(path, exist_ok=exist_ok)
            return
        except Exception as e:
            print(f"[i_o][fs_makedirs][warn] json_proxy.fs_makedirs failed, fallback to os: {e!r}")

    os.makedirs(path, exist_ok=exist_ok)
    print(f"[i_o][fs_makedirs] local mkdir: {path}")

def fs_stat(path: Path):
    """
    S3-aware stat() that returns a stat-like object with st_size and st_mtime attributes.
    """
    if not S3_ENABLED or not _is_s3_path(path):
        return path.stat()
    
    # Try S3 stat first
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "stat"):
            st = _jp.stat(str(path))
            if st is not None:
                return st
    except Exception as e:
        debug(f"[fs_stat] S3 stat failed: {e!r}")
    
    # Fallback: try to get size via read and construct a minimal stat
    try:
        size = fs_stat_size(path)
        class _StatResult:
            st_size = size
            st_mtime = 0
        return _StatResult()
    except Exception:
        pass
    
    # Final fallback
    class _FakeStat:
        st_size = -1
        st_mtime = 0
    return _FakeStat()

def fs_stat_size(path: Path) -> int:
    if not S3_ENABLED or not _is_s3_path(path):
        try:
            return path.stat().st_size
        except Exception:
            return 0
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "stat"):
            st = _jp.stat(str(path))
            if hasattr(st, "st_size"):
                return int(st.st_size)  # type: ignore
            if isinstance(st, dict) and "st_size" in st:
                return int(st["st_size"])
        return path.stat().st_size # Fallback
    except Exception:
        return 0

# ---- Directory creation (no-op on S3) ---------------------------------------

def fs_mkdirs(path: Path, exist_ok: bool = True):
    if not S3_ENABLED or not _is_s3_path(path):
        path.mkdir(parents=True, exist_ok=exist_ok)
        return
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "makedirs"):
            _jp.makedirs(str(path))
    except Exception:
        pass # S3 dirs are virtual, ignore errors
    return

# ---- Listing / walking -------------------------------------------------------

def fs_iterdir(path: Path) -> List[Path]:
    """
    Return children "paths". For S3, we return Path-like shells with joined names.
    """
    if not S3_ENABLED or not _is_s3_path(path):
        try:
            return list(path.iterdir())
        except Exception:
            return []
    
    try:
        from api import json_proxy as _jp
        if hasattr(_jp, "iterdir"):
            items = _jp.iterdir(str(path)) or []
            out: list[Path] = []
            for x in items:
                # json_proxy.iterdir returns Path objects already
                if isinstance(x, Path):
                    out.append(x)
                else:
                    # Fallback just in case
                    out.append(path / str(x))
            return out
        return list(path.iterdir()) # Fallback
    except Exception as e:
        debug(f"[fs_iterdir] S3 call failed: {e!r}")
        return []

# *** END OF FIX 3 REPLACEMENT ***
# --------------------------------

def fs_walk(top: Path) -> Iterator[Tuple[str, List[str], List[str]]]:
    """
    S3-aware replacement for os.walk. Yields (root, dirs, files).
    """
    if not S3_ENABLED or not _is_s3_path(top):
        for root, dirs, files in __import__("os").walk(top):
            yield root, dirs, files
        return

    caps = s3_capabilities()
    if caps["walk"]:
        for root, dirs, files in _s3_call("walk", top, default=[]):
            yield root, dirs, files
        return

    # Fallback: single level
    dirs, files = [], []
    for child in fs_iterdir(top):
        name = Path(child).name
        if fs_is_dir(child):
            dirs.append(name)
        else:
            files.append(name)
    yield str(top), dirs, files

def fs_glob_first(folder: Path, stem: str, allowed_exts: Iterable[str]) -> Optional[Path]:
    """
    Return the first existing file with name {stem}{ext} among allowed_exts.
    """
    for ext in allowed_exts:
        cand = folder / f"{stem}{ext}"
        if fs_exists(cand) and fs_is_file(cand):
            return cand
    return None

# ---- Remove / delete ---------------------------------------------------------

def fs_remove(path: Path) -> bool:
    if not S3_ENABLED or not _is_s3_path(path):
        try:
            path.unlink(missing_ok=True)
            return True
        except Exception as e:
            debug(f"[fs_remove] local failed: {e!r}")
            return False
    caps = s3_capabilities()
    if caps["delete"]:
        ok = _s3_call("delete", path, default=False)
        return bool(ok)
    try:
        write_text(path, "", encoding="utf-8")
        return True
    except Exception:
        return False

def fs_copy(src: Path, dst: Path) -> None:
    """
    S3-aware file copy. Copies a single file from src to dst.
    """
    if not S3_ENABLED or (not _is_s3_path(src) and not _is_s3_path(dst)):
        # Both local - use standard copy
        import shutil
        fs_mkdirs(dst.parent)
        shutil.copy2(src, dst)
        return
    
    # At least one path is S3 - read and write
    try:
        data = fs_read_bytes(src)
        fs_write_bytes(dst, data)
    except Exception as e:
        debug(f"[fs_copy] failed: {e!r}")
        raise

def fs_copytree(src: Path, dst: Path, dirs_exist_ok: bool = True) -> None:
    """
    S3-aware recursive directory copy. Copies entire directory tree from src to dst.
    """
    if not S3_ENABLED or (not _is_s3_path(src) and not _is_s3_path(dst)):
        # Both local - use standard copytree
        import shutil
        shutil.copytree(src, dst, dirs_exist_ok=dirs_exist_ok)
        return
    
    # At least one path is S3 - recursive copy
    try:
        fs_mkdirs(dst)
        for item in fs_iterdir(src):
            item_dst = dst / item.name
            if fs_is_dir(item):
                fs_copytree(item, item_dst, dirs_exist_ok=dirs_exist_ok)
            else:
                fs_copy(item, item_dst)
    except Exception as e:
        debug(f"[fs_copytree] failed: {e!r}")
        raise

def fs_write_bytes(path: Path, data: bytes) -> None:
    """
    S3- and local-aware binary write helper.
    - Creates parent dirs automatically (both local and S3)
    - Uses json_proxy's fs_open_writebin for S3 targets
    - Falls back gracefully to local path.write_bytes()
    """
    try:
        # Always ensure parent directories exist first
        fs_mkdirs(path.parent)

        # If S3 disabled or not an S3 path, just write locally
        if not S3_ENABLED or not _is_s3_path(path):
            path.write_bytes(data)
            return

        # Otherwise, stream directly to S3
        with fs_open_writebin(path) as f:
            f.write(data)
        debug(f"[fs_write_bytes] wrote to S3: {path}")
    except Exception as e:
        debug(f"[fs_write_bytes][error] failed for {path!r}: {e!r}")
        raise


def fs_read_bytes(path: Path) -> bytes:
    """
    S3-aware binary read.
    """
    if not S3_ENABLED or not _is_s3_path(path):
        return path.read_bytes()
    
    try:
        with fs_open_readbin(path) as f:
            return f.read()
    except Exception as e:
        debug(f"[fs_read_bytes] failed: {e!r}")
        raise

def fs_write_text(path: Path, text: str, encoding: str = "utf-8") -> None:
    """
    S3-aware text write (wraps existing write_text from json_proxy).
    """
    fs_mkdirs(path.parent)
    write_text(path, text, encoding=encoding)

def fs_read_text(path: Path, encoding: str = "utf-8") -> str:
    """
    S3-aware text read (wraps existing read_text from json_proxy).
    """
    return read_text(path, encoding=encoding)

def fs_copy(src: Path, dst: Path) -> None:
    """
    S3-aware file copy. Copies a single file from src to dst.
    """
    if not S3_ENABLED or (not _is_s3_path(src) and not _is_s3_path(dst)):
        # Both local - use standard copy
        import shutil
        fs_mkdirs(dst.parent)
        shutil.copy2(src, dst)
        return
    
    # At least one path is S3 - read and write
    try:
        data = fs_read_bytes(src)
        fs_write_bytes(dst, data)
    except Exception as e:
        debug(f"[fs_copy] failed: {e!r}")
        raise

def fs_copytree(src: Path, dst: Path, dirs_exist_ok: bool = True) -> None:
    """
    S3-aware recursive directory copy. Copies entire directory tree from src to dst.
    """
    if not S3_ENABLED or (not _is_s3_path(src) and not _is_s3_path(dst)):
        # Both local - use standard copytree
        import shutil
        shutil.copytree(src, dst, dirs_exist_ok=dirs_exist_ok)
        return
    
    # At least one path is S3 - recursive copy
    try:
        fs_mkdirs(dst)
        for item in fs_iterdir(src):
            item_dst = dst / item.name
            if fs_is_dir(item):
                fs_copytree(item, item_dst, dirs_exist_ok=dirs_exist_ok)
            else:
                fs_copy(item, item_dst)
    except Exception as e:
        debug(f"[fs_copytree] failed: {e!r}")
        raise

# ---- Binary file open helpers ------------------------------------------------

def fs_open_readbin(path: Path):
    """
    Open a readable binary stream for local or S3.
    """
    if not S3_ENABLED or not _is_s3_path(path):
        return builtins.open(path, "rb")
    try:
        from api import json_proxy as _jp  # type: ignore
    except Exception:
        return builtins.open(path, "rb")
    if getattr(_jp, "read_bytes", None):
        data = _jp.read_bytes(path)  # type: ignore[attr-defined]
        return io.BytesIO(data)
    data = read_text(path, encoding="utf-8", errors="ignore")
    return io.BytesIO(data.encode("utf-8"))

def fs_open_writebin(path: Path):
    """
    Open a writable binary stream for local or S3. Upload on close.
    """
    if not S3_ENABLED or not _is_s3_path(path):
        return builtins.open(path, "wb")

    try:
        from api import json_proxy as _jp  # type: ignore
    except Exception:
        return builtins.open(path, "wb")

    has_wb = bool(getattr(_jp, "write_bytes", None))
    buf = io.BytesIO()

    def _close_and_upload():
        body = buf.getvalue()
        if has_wb:
            _jp.write_bytes(path, body)  # type: ignore[attr-defined]
        else:
            write_text(path, body.decode("latin-1"), encoding="latin-1")
        buf.close = lambda: None
        super(io.BytesIO, buf).close()

    buf.close = _close_and_upload  # type: ignore[assignment]
    return buf

# ---- Zip streaming -----------------------------------------------------------

def make_zip_stream(files: Iterable[Tuple[Path, str]]) -> io.BytesIO:
    """
    Create an in-memory ZIP from (path, arcname) pairs.
    S3-aware: reads via fs_open_readbin for each path.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p, arc in files:
            try:
                with fs_open_readbin(p) as fh:
                    zf.writestr(arc, fh.read())
            except Exception as e:
                debug(f"[make_zip_stream] skip {p}: {e!r}")
                continue
    buf.seek(0)
    return buf

# ---- Project listing shim ----------------------------------------------------

def io_list_projects() -> List[str]:
    """
    Return project folder names under the resolved project_root.
    - S3: uses json_proxy.list_projects() if available; else prefixes under root
    - Local: classic iterdir approach (hidden dirs ignored)
    """
    root = resolve_path(Path(), "project_root")

    # S3 mode?
    if S3_ENABLED and _is_s3_path(root):
        caps = s3_capabilities()
        if caps["list_projects"]:
            names = _s3_call("list_projects", default=[]) or []
            return sorted(str(n).rstrip("/").split("/")[-1] for n in names if str(n).strip())
        children = _s3_call("list_dirnames", root, default=[]) or []
        out = []
        for c in children:
            name = str(c).rstrip("/").split("/")[-1]
            if name and not name.startswith("."):
                out.append(name)
        return sorted(set(out))

    try:
        return sorted(p.name for p in root.iterdir() if p.is_dir() and not p.name.startswith("."))
    except Exception:
        return []