# api/routers/noun_workbench/uploads.py
"""CSV/XLSX upload parsing + the (DB-only) bulk preview builder.

Bodies are moved verbatim from the original ``noun_workbench.py``.
"""

import csv
from pathlib import Path
from typing import Any, Dict, List

from core.errors import AppError

from ._router import log
from .validation import (
    _autogen_enforced_blank,
    _generate_primary,
    _list_ids_for_noun,
    validate_item_against_schema,
)


# ──────────────────────────────────────────────────────────────────────────────
# Upload parsing (unchanged)
# ──────────────────────────────────────────────────────────────────────────────
def _parse_upload_to_rows(tmp_path: Path) -> List[Dict[str, Any]]:
    """Parse CSV or XLSX to list[dict]. No pandas dependency.
    Normalizes headers case-insensitively: 'Sample ID' == 'sample id'.
    """
    def normalize_headers(headers: List[str]) -> List[str]:
        seen = {}
        result = []
        for h in headers:
            key = (h or "").strip()
            low = key.lower()
            if low in seen:
                result.append(seen[low])
            else:
                seen[low] = key
                result.append(key)
        return result

    ext = tmp_path.suffix.lower()
    if ext == ".csv":
        with tmp_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise AppError("CSV_MISSING_HEADER", "CSV file missing header row", status=400)
            reader.fieldnames = normalize_headers(reader.fieldnames)
            return [dict(row) for row in reader]

    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(tmp_path, data_only=True)
        sheet = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(sheet.iter_rows(max_row=1))]
        headers = normalize_headers(headers)

        rows = []
        for row in sheet.iter_rows(min_row=2):
            rec = {}
            for h, cell in zip(headers, row):
                rec[h] = "" if cell.value is None else str(cell.value)
            if any(str(v).strip() for v in rec.values()):
                rows.append(rec)
        return rows

    else:
        raise AppError("UNSUPPORTED_FILE_TYPE", f"Unsupported file type: {ext}. Use .csv or .xlsx", status=400)

# ──────────────────────────────────────────────────────────────────────────────
# Preview builder (DB-only)
# ──────────────────────────────────────────────────────────────────────────────
def _normalize_id(val: Any) -> str:
    return str(val or "").strip().lower()

def _preview_rows(
    rows: List[Dict[str, Any]],
    noun_type: str,
    schema: dict,
    project: str,
    project_path: Path,
    mode: str
):
    """
    Return preview structure:
    {valid: [{rowIndex, payload, action}], invalid: [...], warnings: []}
    action ∈ {"update", "insert", "autogen"}
    """
    valid, invalid, warnings = [], [], []
    primary = schema.get("primary_id_field")
    autogen = bool(schema.get("autogenerate_id"))

    existing_ids = set(_list_ids_for_noun(project_path, project, noun_type, primary or "id"))

    for idx, raw in enumerate(rows, start=1):
        payload = {k: ("" if v is None else str(v)) for k, v in raw.items()}

        msg = _autogen_enforced_blank(schema, payload, mode)
        errs = [msg] if msg else []

        errs += validate_item_against_schema(payload, noun_type, schema, project_path, project=project)

        if errs:
            invalid.append({"rowIndex": idx, "payload": payload, "errors": errs})
            continue

        action = None
        pid = payload.get(primary or "")
        if primary and pid:
            norm_pid = _normalize_id(pid)
            action = "update" if norm_pid in {e.lower() for e in existing_ids} else "insert"
        elif autogen:
            action = "autogen"
            try:
                existing = _list_ids_for_noun(project_path, project, noun_type, primary or "id")
            except Exception:
                log.warning("[preview] failed to list existing IDs for autogen", project, noun_type, exc_info=True)
                existing = []
            payload = {**payload, (primary or "id"): _generate_primary(schema, noun_type, project_path, existing)}

        valid.append({"rowIndex": idx, "payload": payload, "action": action})

    return {"valid": valid, "invalid": invalid, "warnings": warnings}
